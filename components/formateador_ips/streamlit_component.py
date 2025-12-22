import streamlit as st
import pandas as pd
from datetime import datetime
from typing import Dict, Optional, Any, cast
import io
from openpyxl.styles import Font, PatternFill, Alignment

# Importar la clase principal
from .ips_formatter import IPSFormatter


def _format_rut_with_dots(rut_number: str, dv: str) -> str:
    rut_digits = ''.join([c for c in str(rut_number) if c.isdigit()])
    rut_digits = rut_digits.lstrip('0')
    if not rut_digits:
        rut_digits = '0'
    dv_clean = str(dv).strip().upper() if dv is not None else ''
    return f"{rut_digits}-{dv_clean}" if dv_clean else rut_digits


def _format_clp_amount(value) -> str:
    try:
        n = int(str(value).strip())
    except Exception:
        n = 0
    return f"${n:,}".replace(',', '.')


def _parse_ips_rendicion_fixed_width(txt_content: str):
    errors = []
    rows = []

    if txt_content is None:
        return pd.DataFrame(), ["Contenido vacío"], []

    lines = txt_content.splitlines()
    for i, raw in enumerate(lines, start=1):
        line = raw.rstrip('\r\n')
        if not line.strip():
            continue

        if len(line) < 83:
            errors.append(f"Línea {i}: longitud {len(line)} < 83")
            line = line.ljust(83)
        elif len(line) > 83:
            errors.append(f"Línea {i}: longitud {len(line)} > 83 (se truncó)")
            line = line[:83]

        record: Dict[str, Any] = {
            'DISA-CODINSC': line[0:2],
            'CODDES': line[2:6],
            'NUMINS': line[6:19],
            'DVNINS': line[19:20],
            'GRUPA': line[20:21],
            'NUMBE': line[21:23],
            'TIPO_PENSIONADO': line[23:24],
            'NOMBRE': line[24:64].rstrip(),
            'RUT': line[64:72],
            'DIG_VERIF': line[72:73],
            'MONTO_DESCUENTO': line[73:83],
        }

        for k, v in list(record.items()):
            if isinstance(v, str):
                record[k] = v.strip() if k != 'NOMBRE' else v

        rut_num = record.get('RUT', '')
        dv = record.get('DIG_VERIF', '')
        record['RUT_COMPLETO'] = _format_rut_with_dots(rut_num, dv)

        monto_raw = ''.join([c for c in str(record.get('MONTO_DESCUENTO', '')).strip() if c.isdigit()])
        monto_int_raw = int(monto_raw) if monto_raw else 0
        monto_int = monto_int_raw // 1000
        record['MONTO_DESCUENTO'] = monto_int
        record['MONTO_FORMATEADO'] = _format_clp_amount(monto_int)

        rows.append(record)

    df = pd.DataFrame(rows)
    return df, errors, lines


def _df_to_styled_excel_bytes(df: pd.DataFrame, sheet_name: str = 'IPS') -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(cast(Any, buffer), engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.book[sheet_name]

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        ws.freeze_panes = "A2"

        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                v = cell.value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    buffer.seek(0)
    return buffer.getvalue()

def generate_ips_template():
    """Genera un archivo XLSX de ejemplo para el formateador IPS"""
    # Datos de ejemplo para el template IPS con todos los campos dinámicos
    data = {
        'RUT': ['12345678-9', '98765432-1', '11111111-1'],
        'NOMBRE': ['JUAN CARLOS PEREZ GONZALEZ', 'MARIA ELENA RODRIGUEZ SILVA', 'CARLOS ANDRES MARTINEZ LOPEZ'],
        'MONTO': [150000, 75000, 200000],
        'CODINSC': [81, 83, 41],
        'NUMINS': ['1234567890123', '9876543210987', '1111111111111'],
        'DVNINS': ['K', '5', '2'],
        'FECINI': ['01/12/2024', '15/01/2025', '01/02/2025'],
        'CANCUO': [1, 12, 6]
    }
    
    df = pd.DataFrame(data)
    
    # Crear archivo Excel en memoria usando ExcelWriter
    buffer = io.BytesIO()
    with pd.ExcelWriter(cast(Any, buffer), engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Datos_IPS')
    buffer.seek(0)
    
    return buffer.getvalue()

def render_ips_formatter():
    """
    Componente Streamlit para el formateador IPS
    Para integrar en tu app existente
    """
    
    st.header("📄 Formateador de Archivos IPS")

    tab_excel_to_txt, tab_txt_to_excel = st.tabs(["Excel → TXT", "TXT → Excel"])

    with tab_excel_to_txt:
        st.markdown("""
        ### 📋 Descripción del Formateador IPS
        
        Esta herramienta te permite convertir archivos Excel con datos de beneficiarios al formato estándar IPS (Instituto de Previsión Social) 
        utilizado para procesar pagos previsionales y beneficios sociales a personas naturales.
        
        ### 📝 Formato del Archivo Excel Requerido
        
        Tu archivo Excel debe contener las siguientes columnas **exactamente con estos nombres**:
        
        | Columna | Descripción | Ejemplo |
        |---------|-------------|---------|
        | **RUT** | RUT del beneficiario sin puntos, solo con guión | 12345678-9 |
        | **NOMBRE** | Nombre completo del beneficiario (máx. 40 caracteres) | JUAN CARLOS PEREZ GONZALEZ |
        | **MONTO** | Monto a transferir (número entero, sin decimales) | 150000 |
        | **CODINSC** | Código de inscripción (1-2 dígitos) | 81 |
        | **NUMINS** | Número de inscripción (13 dígitos) | 1234567890123 |
        | **DVNINS** | Dígito verificador del número de inscripción | K |
        | **FECINI** | Fecha de inicio en formato DD/MM/AAAA | 01/12/2024 |
        | **CANCUO** | Cantidad de cuotas (1-3 dígitos) | 12 |
        
        ### ⚠️ Validaciones Importantes
        - **RUT**: Formato sin puntos, solo con guión (ej: 12345678-9). Debe ser válido según algoritmo chileno
        - **NOMBRE**: No puede estar vacío y máximo 40 caracteres (personas naturales)
        - **MONTO**: Debe ser un número positivo mayor a 0 (máximo 10 dígitos)
        - **CODINSC**: Código numérico de 1-2 dígitos
        - **NUMINS**: Número de inscripción de exactamente 13 dígitos
        - **DVNINS**: Dígito verificador alfanumérico (1 carácter)
        - **FECINI**: Fecha en formato DD/MM/AAAA
        - **CANCUO**: Cantidad de cuotas numérica (1-3 dígitos)
        """)

        excel_template = generate_ips_template()
        st.download_button(
            label="📥 Descargar Template Excel",
            data=excel_template,
            file_name="template_ips.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Descarga un archivo Excel con el formato correcto y ejemplos para el formateador IPS"
        )

        st.markdown("---")

        if 'ips_formatter' not in st.session_state:
            st.session_state.ips_formatter = IPSFormatter()

        formatter = st.session_state.ips_formatter

        with st.sidebar:
            st.subheader("Parámetros Fijos del Archivo")

            tipreg = st.number_input("TIPREG", min_value=0, max_value=9, value=2, help="Tipo de registro")
            atrib = st.number_input("ATRIB", min_value=0, max_value=9, value=0, help="Atributo")
            coddes = st.number_input("CODDES", min_value=1, max_value=9999, value=1005, help="Código de descuento")
            umdesc = st.number_input("UMDESC", min_value=1, max_value=99, value=2, help="Unidad de descuento")
            grupa = st.number_input("GRUPA", min_value=0, max_value=9, value=1, help="Grupo")
            numbe = st.number_input("NUMBE", min_value=1, max_value=99, value=1, help="Número de beneficio")
            numret = st.number_input("NUMRET", min_value=0, max_value=9, value=0, help="Número de retención")
            tipmov = st.number_input("TIPMOV", min_value=0, max_value=9, value=1, help="Tipo de movimiento")

            st.subheader("Fechas y Códigos")

            fecha_mov = st.date_input(
                "Fecha del Movimiento (FECMOV)",
                value=datetime.now().date(),
                help="Mes y año del movimiento (formato MMAAAA)"
            )

            agencia = st.number_input("AGENCIA", min_value=1, max_value=999, value=972, help="Código de agencia")

            fecven_input = st.text_input(
                "FECVEN (Fecha Vencimiento)",
                value="00000000",
                max_chars=8,
                help="Fecha de vencimiento en formato DDMMAAAA (8 dígitos). Por defecto: 00000000"
            )

            if isinstance(fecha_mov, tuple):
                fecha_mov = fecha_mov[0] if fecha_mov else datetime.now().date()

            if fecha_mov:
                mes = fecha_mov.month
                año = fecha_mov.year
            else:
                mes = datetime.now().month
                año = datetime.now().year

        uploaded_file = st.file_uploader(
            "Cargar archivo Excel",
            type=['xlsx', 'xls'],
            help="Archivo Excel con los datos de descuentos"
        )

        if uploaded_file is not None:
            df = pd.read_excel(uploaded_file)

            st.subheader("📊 Vista previa del archivo")

            df_display = df.copy()
            if 'NUMINS' in df_display.columns:
                df_display['NUMINS'] = df_display['NUMINS'].astype(str)

            st.dataframe(df_display.head())

            col1, col2 = st.columns(2)
            with col1:
                st.metric("Filas", len(df))
            with col2:
                st.metric("Columnas", len(df.columns))

            st.subheader("✅ Validación de Columnas")
            required_columns = ['RUT', 'NOMBRE', 'MONTO', 'CODINSC', 'NUMINS', 'DVNINS', 'FECINI', 'CANCUO']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                st.error(f"❌ Faltan las siguientes columnas requeridas: {', '.join(missing_columns)}")
                st.info("💡 Descarga el template Excel para ver el formato correcto")
                return
            else:
                st.success("✅ Todas las columnas requeridas están presentes")

                with st.expander("📋 Columnas detectadas"):
                    for col in required_columns:
                        if col in df.columns:
                            sample_value = df[col].iloc[0] if len(df) > 0 else "N/A"
                            st.write(f"✅ **{col}**: {sample_value}")
                        else:
                            st.write(f"❌ **{col}**: No encontrada")

            if st.button("🔄 Procesar Archivo", type="primary"):
                fixed_params = {
                    'tipreg': tipreg,
                    'atrib': atrib,
                    'coddes': coddes,
                    'umdesc': umdesc,
                    'grupa': grupa,
                    'numbe': numbe,
                    'numret': numret,
                    'tipmov': tipmov,
                    'mes': mes,
                    'año': año,
                    'agencia': agencia,
                    'fecven': fecven_input
                }

                with st.spinner("Procesando archivo..."):
                    content, errors, filename = formatter.process_dataframe_complete(df, fixed_params)

                col1, col2 = st.columns([2, 1])

                with col1:
                    if errors:
                        st.error(f"Se encontraron {len(errors)} errores:")
                        for error in errors[:10]:
                            st.write(f"❌ {error}")

                        if len(errors) > 10:
                            st.write(f"... y {len(errors) - 10} errores más")
                    else:
                        st.success("✅ Archivo procesado sin errores")

                with col2:
                    st.metric("Registros procesados", len(df))
                    st.metric("Errores encontrados", len(errors))

                if not errors or st.checkbox("Mostrar preview aunque haya errores"):
                    with st.expander("📄 Vista previa del archivo generado", expanded=True):
                        content, _, _ = formatter.process_dataframe_complete(df, fixed_params)
                        lines = content.split('\n')

                        pos_line = "".join([str(i % 10) for i in range(1, 117)])
                        st.text("Posiciones de caracteres (1-116):")
                        st.code(pos_line)

                        st.text("Primera línea del archivo generado:")

                        for i, line in enumerate(lines[:1]):
                            if line.strip():
                                st.code(f"Línea {i+1}: {line}")
                                st.text("")
                            else:
                                st.code(f"Línea {i+1}: {line}")

                        st.info("🎨 **Códigos:** 🔵 RUT | 🟢 Datos Excel | 🟡 Datos principales | 🔴 Parámetros fijos | ⚪ Espacios")

                    with st.expander("🔍 Desglose de campos DISA - Línea 1", expanded=False):
                        if lines and lines[0].strip():
                            line = lines[0]
                            padded_line = line.ljust(116)

                            field_info = [
                                ("RUTBEN", padded_line[0:8], "1-8", "🔵"),
                                ("DVRBEN", padded_line[8:9], "9", "🔵"),
                                ("CODINSC", padded_line[9:11], "10-11", "🟢"),
                                ("TIPREG", padded_line[11:12], "12", "🔴"),
                                ("ATRIB", padded_line[12:13], "13", "🔴"),
                                ("CODDES", padded_line[13:17], "14-17", "🔴"),
                                ("UMDESC", padded_line[17:19], "18-19", "🔴"),
                                ("NUMINS", padded_line[19:32], "20-32", "🟢"),
                                ("DVNINS", padded_line[32:33], "33", "🟢"),
                                ("GRUPA", padded_line[33:34], "34", "🔴"),
                                ("NUMBE", padded_line[34:36], "35-36", "🔴"),
                                ("NUMRET", padded_line[36:37], "37", "🔴"),
                                ("TIPMOV", padded_line[37:38], "38", "🔴"),
                                ("NOMBRE", padded_line[38:78], "39-78", "🟡"),
                                ("MONDE", padded_line[78:88], "79-88", "🟡"),
                                ("FECINI", padded_line[88:96], "89-96", "🟢"),
                                ("FECVEN", padded_line[96:104], "97-104", "🔴"),
                                ("CANCUO", padded_line[104:107], "105-107", "🟢"),
                                ("FECMOV", padded_line[107:113], "108-113", "🔴"),
                                ("AGENCIA", padded_line[113:116], "114-116", "🔴")
                            ]

                            st.text("Desglose completo de campos en orden secuencial:")
                            for name, value, pos, emoji in field_info:
                                st.text(f"{emoji} {name}({pos}): '{value}'")

                if content:
                    st.download_button(
                        label="⬇️ Descargar archivo IPS",
                        data=content,
                        file_name=filename,
                        mime="text/plain",
                        help=f"Descargar {filename}"
                    )

                    st.info(f"""
                    📊 **Información del archivo:**
                    - Nombre: `{filename}`
                    - Registros: {len(df)}
                    - Tamaño: {len(content)} bytes
                    - Período: {mes:02d}/{año}
                    """)

    with tab_txt_to_excel:
        st.markdown("""
        ### 📊 Conversor de Archivos IPS a Excel
        
        Sube el archivo TXT de rendición IPS (formato ancho fijo **83 caracteres**) y lo convertiré a un Excel legible.
        """)

        uploaded_txt = st.file_uploader(
            "Cargar archivo TXT IPS",
            type=['txt'],
            help="Archivo TXT IPS de 83 caracteres por registro"
        )

        if uploaded_txt is not None:
            try:
                raw_bytes = uploaded_txt.read()
                try:
                    txt_content = raw_bytes.decode('utf-8')
                except Exception:
                    txt_content = raw_bytes.decode('latin-1')
            except Exception as e:
                st.error(f"No se pudo leer el archivo: {str(e)}")
                return

            df_txt, parse_errors, original_lines = _parse_ips_rendicion_fixed_width(txt_content)

            if parse_errors:
                st.warning(f"Se detectaron {len(parse_errors)} observaciones de parseo")
                with st.expander("Ver detalles"):
                    for msg in parse_errors[:50]:
                        st.write(f"- {msg}")

            if df_txt.empty:
                st.error("No se detectaron registros para convertir")
                return

            df_out = df_txt[[
                'RUT_COMPLETO',
                'NOMBRE',
                'MONTO_FORMATEADO',
                'DISA-CODINSC',
                'CODDES',
                'NUMINS',
                'DVNINS',
                'GRUPA',
                'NUMBE',
                'TIPO_PENSIONADO',
                'RUT',
                'DIG_VERIF',
                'MONTO_DESCUENTO'
            ]].copy()

            st.subheader("📊 Vista previa")
            st.dataframe(df_out.head(50))

            col1, col2 = st.columns(2)
            with col1:
                st.metric("Registros", len(df_out))
            with col2:
                st.metric("Observaciones", len(parse_errors))

            excel_bytes = _df_to_styled_excel_bytes(cast(pd.DataFrame, df_out), sheet_name='IPS')
            st.download_button(
                label="⬇️ Descargar Excel",
                data=excel_bytes,
                file_name="ips_rendicion.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Descarga el Excel generado a partir del TXT IPS"
            )

def ips_formatter_page():
    """
    Página completa para usar como página independiente
    """
    st.set_page_config(
        page_title="Formateador IPS",
        page_icon="📄",
        layout="wide"
    )
    
    render_ips_formatter()

# Función para integrar en una app existente
def add_ips_formatter_to_sidebar():
    """
    Versión compacta para integrar en sidebar de app existente
    """
    with st.sidebar:
        st.subheader("🔧 Formateador IPS")
        
        if st.button("Abrir Formateador IPS"):
            st.session_state.show_ips_formatter = True
    
    if st.session_state.get('show_ips_formatter', False):
        render_ips_formatter()
        
        if st.button("❌ Cerrar Formateador"):
            st.session_state.show_ips_formatter = False

# Para uso como módulo independiente
if __name__ == "__main__":
    ips_formatter_page()